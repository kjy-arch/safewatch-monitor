import smtplib
from io import BytesIO
from datetime import datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.database import supabase
from app.services.excel_safety import sanitize_workbook


# ── 엑셀 생성 ─────────────────────────────────────────────
LEVEL_COLORS = {"높음": "FFCCCC", "중간": "FFF2CC", "낮음": "CCFFCC"}

COLUMNS = [
    ("번호",       5),
    ("출처",       8),
    ("게시일",     14),
    ("거짓점수",   9),
    ("거짓척도",   8),
    ("내용구분",   12),
    ("과목",       10),
    ("분류구분",   18),   # 가이드라인 삭제기준 (006)
    ("조치유형",   10),   # 삭제대상/비대상/종합판단 (006)
    ("의도유형",   10),
    ("내용유형",   10),
    ("판단이유",   30),
    ("소관부서",   14),
    ("소관부서2",  14),   # 복수 부서 매칭 (요구 Q4)
    ("제목",       30),   # 요구 R2가 명시한 항목 — 그동안 누락돼 있었다
    ("원문",       50),
    ("링크",       30),
    ("대응상태",   10),
]


def _build_excel(articles: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "수집결과"

    today_str = datetime.now(timezone(timedelta(hours=9))).strftime("%Y년 %m월 %d일")
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = f"SafeWatch Monitor 수집 결과 — {today_str}"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 헤더
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[2].height = 22

    link_col_idx = next(
        idx for idx, (name, _) in enumerate(COLUMNS, start=1) if name == "링크"
    )

    # 데이터
    for row_num, a in enumerate(articles, start=3):
        def _dname(key):
            v = a.get(key)
            return v.get("name", "") if isinstance(v, dict) else ""
        dept  = _dname("dept1") or _dname("departments")
        dept2 = _dname("dept2")
        pub  = a.get("published_at", "")[:16].replace("T", " ") if a.get("published_at") else ""
        level = a.get("false_level") or "미분류"
        bg    = LEVEL_COLORS.get(level, "FFFFFF")

        row_data = [
            row_num - 2,
            a.get("source_type", ""),
            pub,
            a.get("false_score", ""),
            level,
            a.get("label_l2") or "",
            a.get("subject") or "",
            a.get("category") or "",
            a.get("action_type") or "",
            a.get("intent_type") or "",
            a.get("content_type") or "",
            a.get("false_reason", ""),
            dept,
            dept2,
            (a.get("title") or "")[:200],
            (a.get("content") or "")[:1000],
            a.get("url", ""),
            a.get("response_status", "미확인"),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (9, 10)))
            cell.border    = border
            cell.font      = Font(size=9)
            if col_idx == 3:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            # 열 추가·순서 변경에도 실제 "링크" 열을 찾아 적용한다.
            if (col_idx == link_col_idx and isinstance(value, str)
                    and value.lower().startswith(("http://", "https://"))):
                # 실제 URL을 그대로 보여주면서 셀 자체를 클릭 가능한 링크로 만든다.
                cell.value     = value
                cell.hyperlink = value
                cell.font      = Font(size=9, color="0563C1", underline="single")

    ws.freeze_panes = "A3"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A2:{last_col}{len(articles) + 2}"

    sanitize_workbook(wb)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 통계 요약 ─────────────────────────────────────────────
def _summary_html(articles: list, today_str: str) -> str:
    total   = len(articles)
    high    = sum(1 for a in articles if a.get("false_level") == "높음")
    mid     = sum(1 for a in articles if a.get("false_level") == "중간")
    low     = sum(1 for a in articles if a.get("false_level") == "낮음")
    unclf   = total - high - mid - low

    by_src  = {}
    for a in articles:
        s = a.get("source_type", "기타")
        by_src[s] = by_src.get(s, 0) + 1

    src_rows = "".join(
        f"<tr><td style='padding:4px 12px'>{k}</td><td style='padding:4px 12px;text-align:center'>{v}건</td></tr>"
        for k, v in sorted(by_src.items(), key=lambda x: -x[1])
    )

    return f"""
<html><body style="font-family:'맑은 고딕',sans-serif;color:#333">
<div style="max-width:600px;margin:0 auto">

  <div style="background:#1F4E79;color:white;padding:20px;border-radius:8px 8px 0 0">
    <h2 style="margin:0;font-size:18px">📡 SafeWatch Monitor 일일 수집 결과</h2>
    <p style="margin:4px 0 0;font-size:13px;opacity:.85">{today_str} 수집분 — 첨부 엑셀 파일 참조</p>
  </div>

  <div style="background:#f8f9fa;padding:20px;border:1px solid #dee2e6">

    <h3 style="color:#1F4E79;margin-top:0">📊 오늘의 수집 요약</h3>
    <table style="width:100%;border-collapse:collapse">
      <tr>
        <td style="padding:10px;text-align:center;background:#fff;border-radius:6px;border:1px solid #dee2e6;width:25%">
          <div style="font-size:28px;font-weight:bold;color:#1F4E79">{total}</div>
          <div style="font-size:12px;color:#666">전체 수집</div>
        </td>
        <td style="width:4%"></td>
        <td style="padding:10px;text-align:center;background:#FFCCCC;border-radius:6px;border:1px solid #FFAAAA;width:20%">
          <div style="font-size:24px;font-weight:bold;color:#CC0000">{high}</div>
          <div style="font-size:12px;color:#CC0000">높음</div>
        </td>
        <td style="width:2%"></td>
        <td style="padding:10px;text-align:center;background:#FFF2CC;border-radius:6px;border:1px solid #FFDD99;width:20%">
          <div style="font-size:24px;font-weight:bold;color:#996600">{mid}</div>
          <div style="font-size:12px;color:#996600">중간</div>
        </td>
        <td style="width:2%"></td>
        <td style="padding:10px;text-align:center;background:#CCFFCC;border-radius:6px;border:1px solid #99DD99;width:20%">
          <div style="font-size:24px;font-weight:bold;color:#006600">{low}</div>
          <div style="font-size:12px;color:#006600">낮음</div>
        </td>
      </tr>
    </table>

    <h3 style="color:#1F4E79;margin-top:20px">📂 출처별 수집 현황</h3>
    <table style="width:100%;border-collapse:collapse;background:white;border:1px solid #dee2e6;border-radius:6px">
      {src_rows}
    </table>

    {"<div style='margin-top:16px;padding:12px;background:#FFEEEE;border-left:4px solid #CC0000;border-radius:4px'><strong style='color:#CC0000'>⚠️ 긴급 주의</strong> — 거짓척도 <strong>높음</strong> {high}건이 탐지되었습니다. 첨부 파일의 붉은 행을 우선 확인하세요.</div>" if high > 0 else ""}

    <div style="margin-top:16px;padding:12px;background:#E8F0FB;border-radius:4px;font-size:12px;color:#555">
      📎 첨부 엑셀 파일에 전체 수집 결과가 포함되어 있습니다.<br>
      컬럼: 출처 / 게시일 / 거짓점수 / 척도 / 내용구분 / 과목 / 판단이유 / 소관부서 / 링크 / 대응상태
    </div>
  </div>

  <div style="background:#f1f3f5;padding:10px;text-align:center;font-size:11px;color:#999;border-radius:0 0 8px 8px">
    SafeWatch Monitor · 병무청 정보기획과
  </div>
</div>
</body></html>"""


def _is_alert_candidate(article: dict, min_score: int) -> bool:
    """점수뿐 아니라 최종 조치유형과 2차 검증 결과까지 반영한다.

    2차 확인이 끝난 글은 verify_action을 우선한다. 본문 조회에 실패한 글은 확인되지 않은
    상태이므로 자동 알림에서 제외한다. 아직 2차 확인 전이거나 조회 대상이 아닌 글은 현재
    조치유형을 사용한다.
    """
    if min_score > 0 and (article.get("false_score") or -1) < min_score:
        return False

    verify_status = article.get("verify_status")
    if verify_status == "조회실패":
        return False
    action_type = (
        article.get("verify_action")
        if verify_status == "확인완료"
        else article.get("action_type")
    )
    return action_type == "삭제대상"


# ── 발송 메인 ─────────────────────────────────────────────
def send_alerts():
    """당일 삭제대상 기사를 점수·2차 검증 결과로 필터링해 엑셀 첨부 발송."""
    recipients = (
        supabase.table("alert_settings")
        .select("email, min_score")
        .eq("is_active", True)
        .execute()
        .data
    )
    if not recipients:
        return

    # 오늘 수집분 전체 — 오늘 00:00 KST를 UTC로 변환
    now_kst = datetime.now(timezone(timedelta(hours=9)))
    today_start = now_kst.replace(
        hour=0, minute=0, second=0, microsecond=0
    ).astimezone(timezone.utc)

    articles = (
        supabase.table("crawled_articles")
        .select("*, dept1:departments!crawled_articles_department_id_fkey(name),dept2:departments!crawled_articles_department_id_2_fkey(name)")
        .gte("created_at", today_start.isoformat())
        .order("false_score", desc=True)
        .execute()
        .data
    )

    if not articles:
        print("[알림] 오늘 수집된 기사 없음 — 발송 건너뜀")
        return

    today_str = now_kst.strftime("%Y년 %m월 %d일")
    filename  = f"safewatch_{now_kst.strftime('%Y%m%d')}.xlsx"

    # min_score별로 필터·엑셀을 한 번만 생성 (같은 기준 수신자끼리 재사용)
    # 미분류(false_score 없음) 기사는 min_score > 0 필터에서 제외됨
    cache: dict[int, tuple] = {}
    sent_ids: set = set()
    for r in recipients:
        ms = r.get("min_score") or 0
        if ms not in cache:
            filtered = [a for a in articles if _is_alert_candidate(a, ms)]
            cache[ms] = (
                filtered,
                _build_excel(filtered) if filtered else None,
                _summary_html(filtered, today_str) if filtered else None,
            )
        filtered, excel_bytes, html_body = cache[ms]

        if not filtered:
            print(f"[알림] {r['email']}: 거짓점수 {ms} 이상 기사 없음 — 발송 생략")
            continue

        try:
            if _send(r["email"], f"[SafeWatch] {today_str} 수집 결과",
                     html_body, excel_bytes, filename):
                sent_ids.update(a["id"] for a in filtered)
                print(f"[알림] 발송 완료 → {r['email']} ({len(filtered)}건, 기준점수 {ms})")
        except Exception as e:
            print(f"[알림] 발송 실패 ({r['email']}): {type(e).__name__}: {e}")

    # 발송 완료 표시 — 실제 발송된 메일에 포함된 기사만
    if sent_ids:
        supabase.table("crawled_articles").update({"alert_sent": True}).in_("id", list(sent_ids)).execute()


def _send(to: str, subject: str, html: str, attachment: bytes, filename: str) -> bool:
    if not settings.SMTP_HOST or not settings.SMTP_USER:
        print(f"[알림 미발송] SMTP 미설정 — {to} / {filename} ({len(attachment)//1024}KB)")
        return False

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"]    = settings.SMTP_USER
    msg["To"]      = to
    msg.attach(MIMEText(html, "html", "utf-8"))

    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT) as server:
        server.starttls()
        server.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
        server.send_message(msg)
    return True
