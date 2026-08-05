"""분기 보고서 집계 테스트 — 앞으로 들어올 '실제 수집·분석 결과'가 제대로 담기는지.

Phase 2 이후 분류된 건은 두 축이 모두 채워지므로, 그 형태의 데이터가
요약·피벗에 정확히 반영되는지 확인한다. 복수 부서(요구 Q4) 계상 규칙 포함.

실행: .venv\\Scripts\\python.exe tests\\test_quarterly_report.py
"""
import os, sys
from io import BytesIO

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

for k in ("SUPABASE_URL", "SUPABASE_SECRET_KEY", "GEMINI_API_KEY",
          "NAVER_CLIENT_ID", "NAVER_CLIENT_SECRET", "YOUTUBE_API_KEY"):
    os.environ.setdefault(k, "dummy" if "URL" not in k else "https://dummy.supabase.co")

import openpyxl
from app.services.batch.excel import build_quarterly_excel

failures = []


def check(name, cond, detail=""):
    print(f"  {'OK  ' if cond else 'FAIL'} {name}" + (f" — {detail}" if not cond and detail else ""))
    if not cond:
        failures.append(name)


DEPTS = {"d1": "병역면탈수사과", "d2": "병역판정과", "d3": "홍보과"}

# Phase 2 이후 실제 분류 결과 형태 (두 축 + 복수 부서)
ARTICLES = [
    {"created_at": "2026-08-01T09:00:00Z", "origin": "수집", "source_type": "커뮤니티",
     "category": "편법·속임수·공정성 훼손", "action_type": "삭제대상", "false_level": "높음",
     "intent_type": "악의적 유포", "department_id": "d1", "department_id_2": "d2"},
    {"created_at": "2026-08-02T09:00:00Z", "origin": "수집", "source_type": "언론",
     "category": "정상정보", "action_type": "비대상", "false_level": "낮음",
     "intent_type": "사실 보도", "department_id": "d3", "department_id_2": None},
    {"created_at": "2026-09-03T09:00:00Z", "origin": "업로드", "source_type": "SNS",
     "category": "허위·조작", "action_type": "삭제대상", "false_level": "높음",
     "intent_type": "악의적 유포", "department_id": "d1", "department_id_2": None},
    {"created_at": "2026-09-04T09:00:00Z", "origin": "업로드", "source_type": "유튜브",
     "category": "단순문의·불평", "action_type": "비대상", "false_level": "낮음",
     "intent_type": "단순 오해", "department_id": None, "department_id_2": None},
]

xlsx = build_quarterly_excel(ARTICLES, DEPTS, "2026-08-01 ~ 2026-09-30")
wb = openpyxl.load_workbook(BytesIO(xlsx))

print("[시트 구성]")
for name in ["요약", "월별_분류구분", "부서별_분류구분", "사이트별_분류구분", "부서별_조치유형"]:
    check(name, name in wb.sheetnames)

summary = {}
for r in wb["요약"].iter_rows(min_row=2, values_only=True):
    if r[0] is not None:
        summary[(r[0], r[1])] = r[2]

print("\n[요약 — 총 건수는 글 1건을 1건으로]")
check("총 건수 4", summary.get(("총 건수", None)) == 4, f"got {summary.get(('총 건수', None))}")
check("복수 부서 안내 표기", any(k[0] == "복수 부서 매칭" for k in summary))

print("\n[요약 — 요구사항이 요구한 분포]")
check("조치유형 삭제대상 2", summary.get(("조치유형", "삭제대상")) == 2, f"got {summary.get(('조치유형','삭제대상'))}")
check("조치유형 비대상 2", summary.get(("조치유형", "비대상")) == 2)
check("분류구분 허위·조작 1", summary.get(("분류구분", "허위·조작")) == 1)
check("의도유형 악의적 유포 2", summary.get(("의도유형", "악의적 유포")) == 2)
check("거짓척도 높음 2", summary.get(("거짓척도", "높음")) == 2)
check("출처 SNS 1", summary.get(("출처", "SNS")) == 1)
check("구분 수집 2 / 업로드 2",
      summary.get(("구분", "수집")) == 2 and summary.get(("구분", "업로드")) == 2)

print("\n[부서별 — 복수 부서는 부서마다 계상 (요구 Q4)]")
ws = wb["부서별_분류구분"]
dept_rows = {r[0]: r for r in ws.iter_rows(min_row=2, values_only=True)}
check("병역면탈수사과 있음", "병역면탈수사과" in dept_rows)
check("2순위 부서도 집계됨", "병역판정과" in dept_rows,
      f"부서 목록: {[k for k in dept_rows if k]}")
check("부서 미배정도 표시", "(미배정)" in dept_rows)

total_row = dept_rows.get("합계")
if total_row:
    # 4건 중 1건이 2개 부서 → 부서별 합계는 5
    check("부서별 합계 5 (복수 부서 반영)", total_row[-1] == 5, f"got {total_row[-1]}")

print("\n[월별]")
months = {r[0] for r in wb["월별_분류구분"].iter_rows(min_row=2, values_only=True)}
check("2026-08 집계", "2026-08" in months, f"got {months}")
check("2026-09 집계", "2026-09" in months)

print("\n[빈 데이터]")
wb2 = openpyxl.load_workbook(BytesIO(build_quarterly_excel([], DEPTS, "기간")))
check("데이터 없어도 생성됨", "요약" in wb2.sheetnames)

print()
if failures:
    print(f"FAILED ({len(failures)}): {failures}")
    sys.exit(1)
print("ALL TESTS PASSED")
