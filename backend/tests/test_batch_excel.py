"""배치 엑셀 파싱 테스트 — 원문 열 선택 우선순위 (네트워크·DB 불필요).

실행: .venv\\Scripts\\python.exe tests\\test_batch_excel.py
"""
import asyncio
import os, sys
from io import BytesIO

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

for k in ("SUPABASE_URL", "SUPABASE_SECRET_KEY", "GEMINI_API_KEY",
          "NAVER_CLIENT_ID", "NAVER_CLIENT_SECRET", "YOUTUBE_API_KEY"):
    os.environ.setdefault(k, "dummy" if "URL" not in k else "https://dummy.supabase.co")

import openpyxl
from fastapi import HTTPException, UploadFile
from app.api.upload import upload_excel
from app.services.batch.excel import (
    MAX_DATA_ROWS,
    MAX_UPLOAD_BYTES,
    build_result_excel,
    parse_excel,
)

failures = []


def check(name, cond, detail=""):
    print(f"  {'OK  ' if cond else 'FAIL'} {name}" + (f" — {detail}" if not cond and detail else ""))
    if not cond:
        failures.append(name)


def build(rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


BODY = "지금 172에 53kg인데 5kg만 빼면 공익 가능하냐?"

print("[원문 열 선택 우선순위]")
# 제목·내용이 함께 있으면 반드시 '내용'을 원문으로 — 제목만 분류되면 본문이 통째로 누락된다
rows = parse_excel(build([["제목", "내용", "출처", "링크"],
                          ["공익 문의", BODY, "커뮤니티", "http://x.test/1"]]))
check("제목보다 내용 우선", rows and rows[0]["text"] == BODY, f"got {rows[0]['text'] if rows else None!r}")
check("출처 인식", rows and rows[0]["source_type"] == "커뮤니티")
check("링크 인식", rows and rows[0]["source_url"] == "http://x.test/1")

# '원문'은 '내용'보다도 앞선다 (Monitor 산출 엑셀 형식)
rows = parse_excel(build([["내용", "원문"], ["짧은 요약", BODY]]))
check("내용보다 원문 우선", rows and rows[0]["text"] == BODY, f"got {rows[0]['text'] if rows else None!r}")

# 본문 계열이 없으면 제목이라도 쓴다
rows = parse_excel(build([["제목", "출처"], ["군면제 방법 문의", "커뮤니티"]]))
check("본문 없으면 제목 사용", rows and rows[0]["text"] == "군면제 방법 문의",
      f"got {rows[0]['text'] if rows else None!r}")

print("\n[헤더 탐지]")
# Monitor 산출 엑셀은 1행이 제목 행(병합) — 그 아래 진짜 헤더를 찾아야 한다
rows = parse_excel(build([["SafeWatch Monitor 수집 결과 — 2026년", None, None],
                          ["원문", "출처", "링크"],
                          [BODY, "커뮤니티", "http://x.test/9"]]))
check("상단 제목행 건너뛰기", rows and rows[0]["text"] == BODY, f"got {rows[0]['text'] if rows else None!r}")

print("\n[기타]")
rows = parse_excel(build([["원문"], [BODY], [""], ["   "]]))
check("빈 행 제외", len(rows) == 1, f"got {len(rows)}")

rows = parse_excel(build([["원문", "출처"], [BODY, "이상한출처"]]))
check("알 수 없는 출처는 '언론'으로", rows and rows[0]["source_type"] == "언론",
      f"got {rows[0]['source_type'] if rows else None!r}")

print("\n[업로드 제한]")
try:
    parse_excel(b"")
    check("빈 파일 거부", False, "예외가 발생하지 않음")
except ValueError as e:
    check("빈 파일 거부", "빈 파일" in str(e), str(e))

try:
    parse_excel(b"not an xlsx")
    check("잘못된 형식 거부", False, "예외가 발생하지 않음")
except ValueError as e:
    check("잘못된 형식 거부", ".xlsx" in str(e), str(e))

try:
    parse_excel(b"x" * (MAX_UPLOAD_BYTES + 1))
    check("10MB 초과 거부", False, "예외가 발생하지 않음")
except ValueError as e:
    check("10MB 초과 거부", "10MB" in str(e), str(e))

boundary_rows = [["원문"]] + [[f"본문 {i}"] for i in range(MAX_DATA_ROWS)]
check("1,000행 경계 허용", len(parse_excel(build(boundary_rows))) == MAX_DATA_ROWS)
try:
    parse_excel(build(boundary_rows + [["초과 행"]]))
    check("1,001행 거부", False, "예외가 발생하지 않음")
except ValueError as e:
    check("1,001행 거부", "1,000행" in str(e), str(e))

print("\n[수식 삽입 방지 — 배치 결과]")
xlsx = build_result_excel(
    [{"text": "=HYPERLINK(\"https://evil.test\")", "source_type": "+CMD", "source_url": "@SUM(1,1)"}],
    [{"false_reason": "  -1+1"}],
)
ws = openpyxl.load_workbook(BytesIO(xlsx), data_only=False).active
check("원문을 일반 텍스트로 저장", ws["A2"].data_type != "f" and ws["A2"].value.startswith("'="))
check("출처를 일반 텍스트로 저장", ws["B2"].data_type != "f" and ws["B2"].value.startswith("'+"))
check("URL을 일반 텍스트로 저장", ws["C2"].data_type != "f" and ws["C2"].value.startswith("'@"))
check("앞 공백 뒤 수식 문자도 차단", ws["H2"].data_type != "f" and ws["H2"].value.startswith("'  -"))

print("\n[업로드 API 오류]")


async def call_upload(filename, content):
    return await upload_excel(UploadFile(BytesIO(content), filename=filename))


for filename, content, expected_status, expected_detail in [
    ("legacy.xls", b"x", 400, ".xlsx"),
    ("fake.xlsx", b"not an xlsx", 422, "유효한 .xlsx"),
    ("empty.xlsx", b"", 422, "빈 파일"),
    # 정확히 10MB면 용량 초과(413)가 아니라 다음 형식 검증(422)까지 진행해야 한다.
    ("size-boundary.xlsx", b"x" * MAX_UPLOAD_BYTES, 422, "유효한 .xlsx"),
    ("large.xlsx", b"x" * (MAX_UPLOAD_BYTES + 1), 413, "10MB"),
]:
    try:
        asyncio.run(call_upload(filename, content))
        check(f"{filename} 거부", False, "예외가 발생하지 않음")
    except HTTPException as e:
        check(f"{filename} 거부", e.status_code == expected_status and expected_detail in e.detail,
              f"status={e.status_code}, detail={e.detail}")

print()
if failures:
    print(f"FAILED ({len(failures)}): {failures}")
    sys.exit(1)
print("ALL TESTS PASSED")
